"""
Excel fallback reading with multiple engines and recovery strategies.

This module provides robust Excel reading capabilities with multiple
engine fallbacks, error recovery, and comprehensive error handling
for when standard Excel processing fails.
"""

import logging
import warnings
from pathlib import Path
from typing import Optional, Any, Dict, List, Union, Iterator
from dataclasses import dataclass

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    import pyxlsb
    PYXLSB_AVAILABLE = True
except ImportError:
    PYXLSB_AVAILABLE = False

from .excel_file_repairer import (
    ExcelFileRepairer,
    ExcelProcessingResult,
    ExcelErrorType,
    detect_excel_corruption,
    repair_excel_file,
)

logger = logging.getLogger(__name__)


@dataclass
class EngineCapability:
    """Engine capability description."""
    name: str
    description: str
    supported_formats: List[str]
    read_only: bool = False
    memory_efficient: bool = False


class ExcelFallbackReader:
    """
    Excel reader with multi-engine fallback and recovery capabilities.

    Provides comprehensive Excel reading with automatic fallback through
    multiple engines and recovery mechanisms for corrupted files.
    """

    # Define engine capabilities
    ENGINES = [
        EngineCapability(
            name="calamine",
            description="Fastest, supports corruption recovery",
            supported_formats=[".xlsx", ".xls", ".xlsm", ".xlsb"],
            memory_efficient=True,
        ),
        EngineCapability(
            name="openpyxl",
            description="Standard XLSX support with full feature support",
            supported_formats=[".xlsx", ".xlsm", ".xltx", ".xltm"],
            read_only=True,
        ),
        EngineCapability(
            name="xlrd",
            description="Legacy XLS support",
            supported_formats=[".xls"],
            memory_efficient=True,
        ),
        EngineCapability(
            name="pyxlsb",
            description="Binary XLSB support",
            supported_formats=[".xlsb"],
            memory_efficient=True,
        ),
    ]

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.repairer = ExcelFileRepairer()
        self._available_engines = self._detect_available_engines()

    def _detect_available_engines(self) -> List[str]:
        """
        Detect which Excel reading engines are available.

        Returns:
            List of available engine names
        """
        available = []

        # Check calamine (fastest, if available)
        try:
            import calamine  # noqa: F401
            available.append("calamine")
        except ImportError:
            pass

        # Check openpyxl
        if OPENPYXL_AVAILABLE:
            available.append("openpyxl")

        # Check xlrd
        if XLRD_AVAILABLE:
            available.append("xlrd")

        # Check pyxlsb
        if PYXLSB_AVAILABLE:
            available.append("pyxlsb")

        self.logger.info(f"Available Excel engines: {available}")
        return available

    def _is_engine_available(self, engine_name: str) -> bool:
        """
        Check if a specific engine is available.

        Args:
            engine_name: Name of the engine to check

        Returns:
            True if engine is available
        """
        if engine_name == "calamine":
            try:
                import calamine  # noqa: F401
                return True
            except ImportError:
                return False
        elif engine_name == "openpyxl":
            return OPENPYXL_AVAILABLE
        elif engine_name == "xlrd":
            return XLRD_AVAILABLE
        elif engine_name == "pyxlsb":
            return PYXLSB_AVAILABLE
        else:
            return False

    def _validate_engine_for_file(self, engine_name: str, file_path: Path) -> bool:
        """
        Validate if an engine can handle a specific file format.

        Args:
            engine_name: Name of the engine
            file_path: Path to the Excel file

        Returns:
            True if engine can handle the file format
        """
        file_ext = file_path.suffix.lower()

        for engine in self.ENGINES:
            if engine.name == engine_name:
                return file_ext in engine.supported_formats

        return False

    def _read_with_pandas_engine(
        self,
        file_path: Union[str, Path],
        engine_name: str,
        **kwargs
    ) -> Any:
        """
        Read Excel file using pandas with specific engine.

        Args:
            file_path: Path to the Excel file
            engine_name: Name of the engine to use
            **kwargs: Additional arguments for pandas.read_excel

        Returns:
            DataFrame or ExcelFile object

        Raises:
            ValueError: If engine is not available or fails
        """
        if not PANDAS_AVAILABLE:
            raise ValueError("pandas is not available")

        if engine_name == "calamine" and not self._is_engine_available("calamine"):
            # Try to import calamine and fall back if not available
            try:
                import calamine
            except ImportError:
                raise ValueError("calamine engine not available")

        file_path = Path(file_path)

        # Validate engine compatibility
        if not self._validate_engine_for_file(engine_name, file_path):
            raise ValueError(f"Engine {engine_name} does not support {file_path.suffix} files")

        # Set default parameters for robustness
        read_kwargs = {
            'engine': engine_name,
            'dtype': object,  # Preserve data types
        }
        read_kwargs.update(kwargs)

        self.logger.debug(f"Reading {file_path} with pandas engine: {engine_name}")

        try:
            # First try to get sheet names
            excel_file = pd.ExcelFile(file_path, engine=engine_name)

            # If sheet_name is specified, read that sheet
            sheet_name = read_kwargs.pop('sheet_name', None)
            if sheet_name is not None:
                return pd.read_excel(excel_file, sheet_name=sheet_name, **read_kwargs)
            else:
                return excel_file

        except Exception as e:
            error_msg = f"Pandas engine {engine_name} failed: {str(e)}"
            self.logger.warning(error_msg)
            raise ValueError(error_msg) from e

    def _read_with_openpyxl_direct(
        self,
        file_path: Union[str, Path],
        read_only: bool = True,
        data_only: bool = True,
        **kwargs
    ) -> Any:
        """
        Read Excel file directly with openpyxl.

        Args:
            file_path: Path to the Excel file
            read_only: Whether to open in read-only mode
            data_only: Whether to read cell values instead of formulas
            **kwargs: Additional arguments for openpyxl.load_workbook

        Returns:
            openpyxl Workbook object

        Raises:
            ValueError: If openpyxl is not available or fails
        """
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl is not available")

        file_path = Path(file_path)

        load_kwargs = {
            'read_only': read_only,
            'data_only': data_only,
        }
        load_kwargs.update(kwargs)

        self.logger.debug(f"Reading {file_path} directly with openpyxl")

        try:
            return openpyxl.load_workbook(str(file_path), **load_kwargs)
        except Exception as e:
            error_msg = f"Direct openpyxl reading failed: {str(e)}"
            self.logger.warning(error_msg)
            raise ValueError(error_msg) from e

    def _extract_sheet_data_from_workbook(self, workbook) -> Dict[str, List[List[Any]]]:
        """
        Extract sheet data from openpyxl workbook.

        Args:
            workbook: openpyxl Workbook object

        Returns:
            Dictionary mapping sheet names to data
        """
        sheets_data = {}

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_data = []

            for row in worksheet.iter_rows(values_only=True):
                # Convert all values to strings for consistency
                row_data = [str(cell) if cell is not None else "" for cell in row]
                sheet_data.append(row_data)

            # Remove trailing empty rows
            while sheet_data and all(not cell for cell in sheet_data[-1]):
                sheet_data.pop()

            if sheet_data:  # Only add non-empty sheets
                sheets_data[sheet_name] = sheet_data

        return sheets_data

    def read_excel_with_fallback(
        self,
        file_path: Union[str, Path],
        max_attempts: int = 3,
        **kwargs
    ) -> ExcelProcessingResult:
        """
        Read Excel file with comprehensive fallback mechanisms.

        Args:
            file_path: Path to the Excel file
            max_attempts: Maximum number of repair attempts
            **kwargs: Additional arguments for reading

        Returns:
            ExcelProcessingResult with data or error information
        """
        file_path = Path(file_path)
        result = ExcelProcessingResult(success=False)

        self.logger.info(f"Starting robust Excel reading for: {file_path}")

        # Step 1: Detect corruption indicators
        corruption_indicators = detect_excel_corruption(file_path)
        if corruption_indicators:
            self.logger.warning(f"Corruption indicators detected: {len(corruption_indicators)}")
            for indicator in corruption_indicators:
                self.logger.debug(f"  - {indicator}")

            result.error_type = ExcelErrorType.CORRUPTION
            result.error_message = f"Corruption detected: {'; '.join(corruption_indicators)}"

        # Step 2: Attempt repairs if corruption is detected
        if corruption_indicators and max_attempts > 0:
            self.logger.info("Attempting file repairs...")

            # Attempt ZIP structure repair
            if "Invalid specification" in " ".join(corruption_indicators) or "Absolute file paths" in " ".join(corruption_indicators):
                try:
                    success, message = repair_excel_file(file_path)
                    result.repair_attempts.append(f"ZIP repair: {message}")
                    if success:
                        self.logger.info("ZIP repair completed successfully")
                    else:
                        self.logger.warning(f"ZIP repair failed: {message}")
                except Exception as e:
                    error_msg = f"ZIP repair failed: {str(e)}"
                    result.repair_attempts.append(error_msg)
                    self.logger.error(error_msg)

        # Step 3: Try each available engine
        engines_to_try = [e for e in self.ENGINES if e.name in self._available_engines]

        for engine in engines_to_try:
            if not self._validate_engine_for_file(engine.name, file_path):
                self.logger.debug(f"Skipping {engine.name} - incompatible format")
                continue

            self.logger.info(f"Trying engine: {engine.name} ({engine.description})")

            try:
                if engine.name == "calamine":
                    try:
                        data = self._read_with_pandas_engine(file_path, engine.name, **kwargs)
                        result.success = True
                        result.data = data
                        result.fallback_used = engine.name
                        self.logger.info(f"Successfully read with {engine.name}")
                        break
                    except ImportError:
                        self.logger.warning("calamine not available, skipping")
                        continue

                elif engine.name == "openpyxl":
                    # Try pandas first, then direct openpyxl
                    try:
                        if PANDAS_AVAILABLE:
                            data = self._read_with_pandas_engine(file_path, engine.name, **kwargs)
                            if hasattr(data, 'sheet_names'):
                                # This is an ExcelFile object
                                sheets_data = {}
                                for sheet_name in data.sheet_names:
                                    try:
                                        df = pd.read_excel(data, sheet_name=sheet_name)
                                        sheets_data[sheet_name] = df.values.tolist()
                                    except Exception:
                                        sheets_data[sheet_name] = []
                                data = sheets_data
                        else:
                            # Direct openpyxl reading
                            workbook = self._read_with_openpyxl_direct(file_path)
                            data = self._extract_sheet_data_from_workbook(workbook)
                            workbook.close()

                        # Check if we actually got meaningful data
                        if hasattr(data, 'sheet_names') and len(data.sheet_names) == 0:
                            # Found 0 sheets, treat as failure to trigger raw extraction
                            result.success = False
                            result.error_type = ExcelErrorType.MISSING_SHEETS
                            result.error_message = f"Engine {engine.name} found 0 sheets"
                            self.logger.warning(f"Engine {engine.name} found 0 sheets, treating as failure")
                        else:
                            result.success = True
                            result.data = data
                            result.fallback_used = engine.name
                            result.sheets_detected = len(data.sheet_names) if hasattr(data, 'sheet_names') else 0
                            self.logger.info(f"Successfully read with {engine.name}")
                        break

                    except Exception as e:
                        self.logger.warning(f"pandas + {engine.name} failed: {str(e)}")

                        # Try direct openpyxl as fallback
                        try:
                            workbook = self._read_with_openpyxl_direct(file_path)
                            data = self._extract_sheet_data_from_workbook(workbook)
                            workbook.close()

                            if data:  # Check if we got any data
                                result.success = True
                                result.data = data
                                result.fallback_used = f"{engine.name} (direct)"
                                self.logger.info(f"Successfully read with direct {engine.name}")
                                break
                        except Exception as e2:
                            self.logger.warning(f"Direct {engine.name} also failed: {str(e2)}")

                elif engine.name == "xlrd" and PANDAS_AVAILABLE:
                    data = self._read_with_pandas_engine(file_path, engine.name, **kwargs)
                    result.success = True
                    result.data = data
                    result.fallback_used = engine.name
                    self.logger.info(f"Successfully read with {engine.name}")
                    break

                elif engine.name == "pyxlsb" and PANDAS_AVAILABLE:
                    data = self._read_with_pandas_engine(file_path, engine.name, **kwargs)
                    result.success = True
                    result.data = data
                    result.fallback_used = engine.name
                    self.logger.info(f"Successfully read with {engine.name}")
                    break

            except Exception as e:
                error_msg = f"Engine {engine.name} failed: {str(e)}"
                result.repair_attempts.append(error_msg)
                self.logger.warning(error_msg)
                continue

        # Step 4: If all engines failed or returned 0 sheets, try raw content extraction
        if not result.success or result.sheets_detected == 0:
            self.logger.info("Standard engines found no sheets, attempting raw content extraction...")

            try:
                raw_data = self.repairer.extract_raw_worksheet_content(file_path)
                if raw_data:
                    result.success = True
                    result.data = raw_data
                    result.fallback_used = "raw extraction"
                    result.sheets_detected = len(raw_data)
                    result.repair_attempts.append("Raw content extraction successful")
                    self.logger.info(f"Raw content extraction succeeded - found {len(raw_data)} sheets")
                else:
                    result.repair_attempts.append("Raw content extraction failed - no data found")
                    self.logger.warning("Raw content extraction failed")
            except Exception as e:
                error_msg = f"Raw extraction failed: {str(e)}"
                result.repair_attempts.append(error_msg)
                self.logger.error(error_msg)

        # Step 5: Final result processing
        if result.success:
            # Count sheets detected
            if isinstance(result.data, dict):
                result.sheets_detected = len(result.data)
            elif hasattr(result.data, 'sheet_names'):
                result.sheets_detected = len(result.data.sheet_names)
            elif hasattr(result.data, 'keys'):
                result.sheets_detected = len(list(result.data.keys()))
            else:
                result.sheets_detected = 1

            self.logger.info(f"Excel reading successful - {result.sheets_detected} sheets detected")

            if result.repair_attempts:
                result.error_message = f"Success after {len(result.repair_attempts)} repair attempts"
            else:
                result.error_message = "Success with standard reading"
        else:
            self.logger.error("All Excel reading attempts failed")

            if not result.error_message:
                result.error_message = "All reading engines and recovery methods failed"

        return result

    def get_available_engines(self) -> List[str]:
        """
        Get list of available engine names.

        Returns:
            List of available engine names
        """
        return [engine.name for engine in self.get_engine_capabilities()]

    def get_engine_capabilities(self) -> List[EngineCapability]:
        """
        Get capabilities of all engines.

        Returns:
            List of EngineCapability objects
        """
        available_capabilities = []
        for engine in self.ENGINES:
            if self._is_engine_available(engine.name):
                available_capabilities.append(engine)

        return available_capabilities

    def test_file_compatibility(self, file_path: Union[str, Path]) -> Dict[str, bool]:
        """
        Test file compatibility with available engines.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary mapping engine names to compatibility status
        """
        file_path = Path(file_path)
        compatibility = {}

        for engine in self.ENGINES:
            if self._is_engine_available(engine.name):
                compatibility[engine.name] = self._validate_engine_for_file(engine.name, file_path)
            else:
                compatibility[engine.name] = False

        return compatibility


def read_excel_with_fallback(file_path: Union[str, Path], **kwargs) -> ExcelProcessingResult:
    """
    Convenience function to read Excel with fallback mechanisms.

    Args:
        file_path: Path to the Excel file
        **kwargs: Additional arguments for reading

    Returns:
        ExcelProcessingResult with data or error information
    """
    reader = ExcelFallbackReader()
    return reader.read_excel_with_fallback(file_path, **kwargs)


def get_available_excel_engines() -> List[str]:
    """
    Get list of available Excel reading engines.

    Returns:
        List of available engine names
    """
    reader = ExcelFallbackReader()
    return [engine.name for engine in reader.get_engine_capabilities()]